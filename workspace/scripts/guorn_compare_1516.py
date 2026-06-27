# -*- coding: utf-8 -*-
import openpyxl
from pathlib import Path

base = Path(r"E:\量化系统\Knowledge\果仁回测结果")
# sheet order: 0=收益统计 3=年度收益统计 9=交易段持仓清单 10=历史交易记录

def grab(f):
    wb = openpyxl.load_workbook(base / (f + ".xlsx"), read_only=True, data_only=True)
    sh = wb.worksheets
    out = {}
    out["收益"] = [list(r) for r in sh[0].iter_rows(min_row=2, max_row=2, values_only=True)][0][:6]
    out["年度"] = [list(r)[:3] for r in sh[3].iter_rows(min_row=2, max_row=5, values_only=True)]
    out["持仓"] = [list(r)[:5] for r in sh[9].iter_rows(min_row=2, max_row=5, values_only=True)]
    wb.close()
    return out

a = grab("15_sm_微盘基准50")
b = grab("16_sm_noc_纯市值正盈利_v4")
print("收益统计 identical:", a["收益"] == b["收益"])
print("年度收益 identical:", a["年度"] == b["年度"])
print("持仓清单 identical:", a["持仓"] == b["持仓"])
print("15 持仓 row2:", a["持仓"][0])
print("16 持仓 row2:", b["持仓"][0])
print("15 年度 row2:", a["年度"][0])
print("16 年度 row2:", b["年度"][0])
