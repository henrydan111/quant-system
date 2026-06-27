# -*- coding: utf-8 -*-
"""Content-based duplicate check across all 65 backtest xlsx.
Two DIFFERENT strategies sharing identical 收益统计 + 年度 + 持仓 content => a
stale/mis-captured export (the exec page exported the previous strategy's result).
"""
import re
import hashlib
from collections import defaultdict
from pathlib import Path
import openpyxl

base = Path(r"E:\量化系统\Knowledge\果仁回测结果")

def content_sig(p):
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    sh = wb.worksheets
    parts = []
    parts.append(tuple([list(r) for r in sh[0].iter_rows(min_row=2, max_row=2, values_only=True)][0][:9]))  # 本策略 row
    parts.append(tuple(tuple(list(r)[:3]) for r in sh[3].iter_rows(min_row=2, max_row=8, values_only=True)))  # 年度
    parts.append(tuple(tuple(list(r)[:5]) for r in sh[9].iter_rows(min_row=2, max_row=8, values_only=True)))  # 持仓
    wb.close()
    return hashlib.md5(repr(parts).encode("utf-8")).hexdigest(), parts[0]

groups = defaultdict(list)
metrics = {}
for p in sorted(base.glob("[0-9][0-9]_*.xlsx")):
    sig, strat_row = content_sig(p)
    groups[sig].append(p.name)
    metrics[p.name] = strat_row

dups = {k: v for k, v in groups.items() if len(v) > 1}
print(f"checked {sum(len(v) for v in groups.values())} files; {len(dups)} duplicate-content group(s)")
for k, v in dups.items():
    print("  DUP:", v, " | 本策略总收益=", metrics[v[0]][0])
if not dups:
    print("NO content duplicates - all 65 are distinct backtests.")
